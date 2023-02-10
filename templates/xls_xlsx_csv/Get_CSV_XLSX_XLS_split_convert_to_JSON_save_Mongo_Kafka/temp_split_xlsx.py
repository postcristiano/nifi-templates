import os, sys
import pandas as pd
from openpyxl import load_workbook, Workbook


def split_xlsx(self, xlsxFiles, dir):
    for file_path in xlsxFiles:
        file_path_dir = os.path.join(dir, file_path)
        wb = load_workbook(filename=file_path_dir)
        fname = os.path.basename(file_path_dir)
        fname = os.path.splitext(fname)[0]
        for sheet in wb.worksheets:
            new_wb = Workbook()
            ws = new_wb.active
            for row_data in sheet.iter_rows():
                for row_cell in row_data:
                    ws[row_cell.coordinate].value = row_cell.value
            new_wb.save(dir + self.split_dir +
                        "{}_{}.xlsx".format(fname, sheet.title))
        os.remove(file_path_dir)


class Temp():
    def __init__(self):
        self.split_dir = "split_xlsx/"

    def format_files(self, dir):
        try:
            sys.stdin.readlines()
            xlsxFiles = sys.stdin
            if not os.path.exists(dir + self.split_dir):
                os.makedirs(dir + self.split_dir)
            xlsxFiles = [
                file for file in os.listdir(dir) if file.endswith('.xlsx')
            ]
            split_xlsx(self, xlsxFiles, dir)
            sys.stdout.write("Script execution finished successfully")
        except Exception as e:
            print("Exception occurred as {}".format(e))
            return False

fa = Temp()
fa.format_files("/path/to/folder")

